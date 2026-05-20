import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule

wb = Workbook()

# ============ Styles ============
thin = Side(border_style='thin', color='B0B0B0')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
F_TITLE = Font(name='標楷體', size=18, bold=True, color='FFFFFF')
F_H2    = Font(name='標楷體', size=12, bold=True, color='FFFFFF')
F_BODY  = Font(name='標楷體', size=10)
F_BOLD  = Font(name='標楷體', size=11, bold=True)
FILL_TITLE = PatternFill('solid', start_color='1F3864')
FILL_H2    = PatternFill('solid', start_color='4472C4')
FILL_WARN  = PatternFill('solid', start_color='FFF2CC')
FILL_OK    = PatternFill('solid', start_color='C6EFCE')
FILL_BAD   = PatternFill('solid', start_color='FFC7CE')

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_H2
        c.fill = FILL_H2
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 32
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def merge_title(ws, text, cols, row=1, fill=FILL_TITLE, font=F_TITLE, height=30):
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = font
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = height

# ============ Sheet: 儀表板 ============
dash = wb.active
dash.title = '儀表板'

merge_title(dash, '🎯 公傳室 J 類廣告追蹤儀表板  |  2026臺中長照台日論壇（案號115B17）', 6, row=1, height=34)
dash.cell(row=2, column=1, value='論壇日：2026/7/3（五）　|　今日請填：').font = Font(name='標楷體', size=10)
dash.cell(row=2, column=2, value='=TODAY()')
dash.cell(row=2, column=2).number_format = 'yyyy/m/d'
dash.cell(row=2, column=3, value='　|　距論壇剩餘天數：').font = Font(name='標楷體', size=10)
dash.cell(row=2, column=4, value='=DATE(2026,7,3)-TODAY()')
dash.cell(row=2, column=4).font = Font(name='標楷體', size=11, bold=True, color='C00000')

# KPI table
header_row(dash, 4,
    ['項目','標規要求','實際','差距','風險/扣款試算','狀態'],
    widths=[28, 22, 12, 12, 30, 14])

# Rows: link to each sheet's totals
# We'll use rows 5-8 for J-1, J-3 (qty), J-3 (reach), J-4, J-5
kpi = [
    # (label, requirement_text, actual_formula, gap_formula, deduction_formula, status_formula)
    ('J-1 報紙廣告刊登則數', '≥2則', "=COUNTIF('J-1報紙廣告'!G5:G14,\"已刊登\")",
        "=MAX(0,2-C5)", "=D5*20000",
        "=IF(C5>=2,\"✓達標\",\"⚠未達 扣\"&TEXT(D5*20000,\"#,##0\")&\"元\")"),
    ('J-3 數位廣編稿則數', '≥10則', "=COUNTIF('J-3數位廣編稿'!J5:J14,\"已發布\")",
        "=MAX(0,10-C6)", "=D6*10000",
        "=IF(C6>=10,\"✓達標\",\"⚠未達 扣\"&TEXT(D6*10000,\"#,##0\")&\"元\")"),
    ('J-3 廣編稿總曝光人次', '≥2,000,000', "=SUM('J-3數位廣編稿'!H5:H14)",
        "=MAX(0,2000000-C7)", "\"以契約價金20%為限\"",
        "=IF(C7>=2000000,\"✓達標\",\"⚠未達200萬\")"),
    ('J-3 長照3.0+失智專題', '≥2篇', "=COUNTIFS('J-3數位廣編稿'!B5:B14,\"⭐政策+失智專題\",'J-3數位廣編稿'!J5:J14,\"已發布\")",
        "=MAX(0,2-C8)", "\"併入J-3則數扣款\"",
        "=IF(C8>=2,\"✓達標\",\"⚠未達\")"),
    ('J-4 電視牆/看板檔次', '≥4,000檔次', "=SUM('J-4電視牆看板'!E5:E14)",
        "=MAX(0,4000-C9)", "=D9*1000",
        "=IF(C9>=4000,\"✓達標\",\"⚠每檔次扣1,000\")"),
    ('J-4 場域數', '≥2處', "=COUNTA('J-4電視牆看板'!B5:B14)",
        "=MAX(0,2-C10)", "\"併入檔次扣款\"",
        "=IF(C10>=2,\"✓達標\",\"⚠未達\")"),
    ('J-5 網路關鍵字廣告', '≥20則', "=COUNTIF('J-5網路關鍵字'!H5:H34,\"已上架\")",
        "=MAX(0,20-C11)", "=D11*1000",
        "=IF(C11>=20,\"✓達標\",\"⚠每則扣1,000\")"),
    ('J-2 論壇後報紙新聞稿', '≥1則', "=COUNTIF('J-2活動後新聞稿'!E5:E9,\"已刊登\")",
        "=MAX(0,1-C12)", "=D12*10000",
        "=IF(C12>=1,\"✓達標\",\"⚠未達 扣10,000\")"),
]
for i, (lbl, req, act, gap, ded, st) in enumerate(kpi, start=5):
    dash.cell(row=i, column=1, value=lbl).font = F_BOLD
    dash.cell(row=i, column=2, value=req)
    dash.cell(row=i, column=3, value=act)
    dash.cell(row=i, column=4, value=gap)
    dash.cell(row=i, column=5, value=ded)
    dash.cell(row=i, column=6, value=st)
    for c in range(1, 7):
        cell = dash.cell(row=i, column=c)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        if c != 1:
            cell.font = F_BODY
        if c == 3:
            cell.number_format = '#,##0'
    dash.row_dimensions[i].height = 28

# Conditional format on status column
status_range = f'F5:F{4+len(kpi)}'
dash.conditional_formatting.add(status_range,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("✓",F5))'], fill=FILL_OK))
dash.conditional_formatting.add(status_range,
    FormulaRule(formula=[f'ISNUMBER(SEARCH("⚠",F5))'], fill=FILL_BAD))

# Summary
row = 5 + len(kpi) + 1
dash.cell(row=row, column=1, value='💰 累計扣款風險（試算）').font = F_BOLD
dash.cell(row=row, column=2, value=f'=D5*20000+D6*10000+D9*1000+D11*1000+D12*10000')
dash.cell(row=row, column=2).number_format = '"NT$"#,##0'
dash.cell(row=row, column=2).font = Font(name='標楷體', size=14, bold=True, color='C00000')
dash.cell(row=row, column=3, value='元（不含200萬人次未達之20%契約價金扣減）').font = Font(name='標楷體', size=9, color='666666')
dash.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
for c in range(1, 7):
    dash.cell(row=row, column=c).fill = FILL_WARN
    dash.cell(row=row, column=c).border = border
dash.row_dimensions[row].height = 30

# Critical dates reminder
row += 2
dash.cell(row=row, column=1, value='⏰ 關鍵截止日').font = F_BOLD
dash.cell(row=row, column=1).fill = FILL_H2
dash.cell(row=row, column=1).font = Font(name='標楷體', size=11, bold=True, color='FFFFFF')
dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
dates = [
    ('2026/6/3', '⚠ 最早截止', 'J-5 網路關鍵字廣告 ≥20則（論壇前30天）'),
    ('2026/6/15', '', 'B-1 主視覺／B-2 海報／H-1~H-4 文宣設計'),
    ('2026/6/19', '⚠ 多項並列', 'J-1 報紙廣告 ≥2則／J-3 廣編稿 ≥10則／J-4 電視牆 ≥4,000檔'),
    ('2026/7/2',  '', 'H-3 識別證/資料袋實體送達會場'),
    ('2026/7/3',  '🎯 論壇日', 'L-1 照片+新聞稿／L-2 成果影片現場剪輯'),
    ('2026/7/3後','', 'J-2 論壇後報紙新聞稿 ≥1篇'),
    ('2026/7/19', '', 'M-1 長照系列影片 ≥10分鐘'),
]
row += 1
for d, mark, desc in dates:
    dash.cell(row=row, column=1, value=d).font = F_BOLD
    dash.cell(row=row, column=2, value=mark).font = Font(name='標楷體', size=10, color='C00000', bold=True)
    dash.cell(row=row, column=3, value=desc).font = F_BODY
    dash.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    for c in range(1, 7):
        dash.cell(row=row, column=c).border = border
        dash.cell(row=row, column=c).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    row += 1

dash.freeze_panes = 'A5'

# ============ J-1 報紙廣告 ============
ws = wb.create_sheet('J-1報紙廣告')
merge_title(ws, 'J-1 報紙廣告追蹤表  |  目標：≥2則 / 三大報 / 截止 2026/6/19  |  扣款：每不足1則扣NT$20,000', 9)
ws.cell(row=2, column=1, value='大報範圍：中國時報、聯合報、自由時報；建議刊在「健康/醫療版」或「中部版」').font = Font(name='標楷體', size=10, color='666666')
ws.merge_cells('A2:I2')
header_row(ws, 4,
    ['序號','報別','版面','刊登標題','刊期(日)','聯絡窗口','狀態','刊登截圖檔名','備註'],
    widths=[6, 14, 14, 36, 12, 14, 12, 22, 22])
samples = [
    (1,'聯合報','中部新聞版','（待擬）長照3.0台中啟動 失智照護升級','','','待安排','','含H-1主視覺'),
    (2,'自由時報','健康醫療版','（待擬）2026臺中長照台日論壇預告','','','待安排','','論壇前1週內'),
]
for r, row in enumerate(samples, start=5):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = F_BODY
        cell.alignment = Alignment(horizontal='center' if c in (1,2,3,5,7) else 'left', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[r].height = 28
# add blank rows
for r in range(7, 15):
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c, value=(r-4) if c==1 else '')
        cell.font = F_BODY
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 26
ws.freeze_panes = 'A5'

# ============ J-2 論壇後新聞稿 ============
ws = wb.create_sheet('J-2活動後新聞稿')
merge_title(ws, 'J-2 論壇後報紙新聞稿  |  目標：≥1篇 / 論壇後儘速 / 扣款：未達扣NT$10,000', 6)
header_row(ws, 4,
    ['序號','報別','刊登標題','刊期','狀態','備註'],
    widths=[6, 14, 40, 12, 12, 30])
for r in range(5, 10):
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c, value=(r-4) if c==1 else '')
        cell.font = F_BODY
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 28
ws.freeze_panes = 'A5'

# ============ J-3 廣編稿（同前版） ============
ws = wb.create_sheet('J-3數位廣編稿')
merge_title(ws, 'J-3 數位廣編稿追蹤表  |  目標：≥10則 / ≥2篇政策+失智專題 / ≥200萬人次 / 截止 2026/6/19', 11)
ws.cell(row=2, column=1, value='扣款：每不足1則扣NT$10,000；未達200萬人次以契約價金20%為限').font = Font(name='標楷體', size=10, color='C00000', bold=True)
ws.merge_cells('A2:K2')
header_row(ws, 4,
    ['序號','類別標記','廣編稿標題','建議媒體/平台','負責窗口','發布日期','URL連結','觸及/曝光數','截圖留存','狀態','備註'],
    widths=[6, 18, 42, 28, 12, 12, 28, 14, 12, 10, 22])
data = [
    (1,'⭐政策+失智專題','長照3.0上路！台中接軌中央八大目標 失智照護全面升級','聯合新聞網元氣網／1966長照粉專'),
    (2,'⭐政策+失智專題','長照3.0新制 失智症族群全納入！台中失智共照網守護2.5萬名家屬','台灣失智症協會粉專／健康2.0'),
    (3,'在地安老','全國最大長照網！台中ABC據點1,879處 三里一站照顧好厝邊','愛長照／家天使'),
    (4,'醫照整合','穩坐六都第一 台中「在宅醫療＋共融照護」打造全民照護網','工商時報網／優照護'),
    (5,'智慧長照','長照即時通App上線 大數據預測需求 台中邁向智慧長照','數位時代／智慧樂齡網'),
    (6,'機構量能','台中首家市立綜合長照機構落成 慈濟攜手打造智慧整合照護','聯合報／大愛新聞'),
    (7,'失智友善','失智友善社區動起來！台中商家化身「友善天使」 走失不漏接','台中市衛生局粉專／失智症協會'),
    (8,'偏鄉長照','偏鄉不孤單 台中山海線遠距醫療＋IDS計畫 補齊長照最後一哩','中山醫附設粉專／自由時報中部'),
    (9,'健康老化','延緩失能從預防開始 台中社區據點打造高齡健康生活圈','樂齡網／50+好好'),
    (10,'國際交流','台日交流 共學失智照護！2026臺中長照台日論壇7/3登場','台中市政府粉專／中時新聞網'),
]
for r, (no, cat, title, media) in enumerate(data, start=5):
    row_vals = [no, cat, title, media, '', '', '', '', '', '待發布', '']
    for c, v in enumerate(row_vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = F_BODY
        cell.alignment = Alignment(horizontal='center' if c in (1,2,6,8,10) else 'left', vertical='center', wrap_text=True)
        cell.border = border
        if c == 8:
            cell.number_format = '#,##0'
    ws.row_dimensions[r].height = 36
ws.freeze_panes = 'D5'

# ============ J-4 電視牆/看板 ============
ws = wb.create_sheet('J-4電視牆看板')
merge_title(ws, 'J-4 電視牆/電子看板追蹤表  |  目標：≥2場域 / 合計≥4,000檔次 / 截止 2026/6/19  |  扣款：每不足1檔扣NT$1,000', 7)
header_row(ws, 4,
    ['序號','場域名稱','地點/地址','刊播期間','檔次','驗收憑證','備註'],
    widths=[6, 24, 28, 18, 12, 22, 26])
for r in range(5, 15):
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c, value=(r-4) if c==1 else '')
        cell.font = F_BODY
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if c == 5:
            cell.number_format = '#,##0'
    ws.row_dimensions[r].height = 28
# Total row
ws.cell(row=15, column=4, value='合計檔次').font = F_BOLD
ws.cell(row=15, column=5, value='=SUM(E5:E14)').font = F_BOLD
ws.cell(row=15, column=5).number_format = '#,##0'
ws.cell(row=15, column=5).fill = FILL_WARN
for c in range(1, 8):
    ws.cell(row=15, column=c).border = border
ws.freeze_panes = 'A5'

# ============ J-5 網路關鍵字 ============
ws = wb.create_sheet('J-5網路關鍵字')
merge_title(ws, 'J-5 網路關鍵字/社群廣告追蹤表  |  目標：≥20則 / 截止 2026/6/3（論壇前30天，⚠最早截止）', 9)
ws.cell(row=2, column=1, value='平台範圍：Yahoo、Google、Facebook、YouTube、Instagram；扣款：每不足1則扣NT$1,000').font = Font(name='標楷體', size=10, color='C00000', bold=True)
ws.merge_cells('A2:I2')
header_row(ws, 4,
    ['序號','平台','廣告類型','關鍵字/標題','投放期間','預算','觸及/點擊','狀態','備註'],
    widths=[6, 14, 16, 28, 18, 12, 14, 12, 22])
preset = [
    ('Google Ads','關鍵字搜尋','長照3.0'),
    ('Google Ads','關鍵字搜尋','失智症照護'),
    ('Google Ads','關鍵字搜尋','台中長照'),
    ('Google Ads','關鍵字搜尋','台日長照論壇'),
    ('Yahoo','關鍵字搜尋','長照3.0'),
    ('Yahoo','關鍵字搜尋','失智共照中心'),
    ('Facebook','貼文加強推廣','論壇預告主視覺貼文'),
    ('Facebook','貼文加強推廣','長照3.0八大目標懶人包'),
    ('Facebook','貼文加強推廣','失智友善天使招募'),
    ('Facebook','貼文加強推廣','長照即時通App介紹'),
    ('Instagram','Reels/Story','論壇預告短影音'),
    ('Instagram','Reels/Story','失智症關懷短影音'),
    ('Instagram','貼文廣告','ABC據點懶人包圖卡'),
    ('YouTube','TrueView前導','論壇15秒預告片'),
    ('YouTube','TrueView前導','長照3.0說明影片'),
    ('YouTube','Bumper Ads','失智友善6秒CTA'),
    ('LINE','LAP原生','長照3.0新制提醒'),
    ('Dcard','社群廣告','年輕族群-照顧者議題'),
    ('Threads','貼文推廣','論壇預告'),
    ('電子報','EDM','合作媒體電子報×1'),
]
for r, (plat, typ, kw) in enumerate(preset, start=5):
    row_vals = [r-4, plat, typ, kw, '', '', '', '待上架', '']
    for c, v in enumerate(row_vals, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = F_BODY
        cell.alignment = Alignment(horizontal='center' if c in (1,2,3,6,7,8) else 'left', vertical='center', wrap_text=True)
        cell.border = border
        if c in (6, 7):
            cell.number_format = '#,##0'
    ws.row_dimensions[r].height = 26
ws.freeze_panes = 'A5'

# ============ 扣款風險清單 ============
ws = wb.create_sheet('扣款風險清單')
merge_title(ws, '⚠ 扣款風險清單 / 公傳室負責範疇', 7, fill=PatternFill('solid', start_color='C00000'))
header_row(ws, 3,
    ['序號','扣款項目','標規依據','扣款計算','上限','驗收憑證','對應分頁'],
    widths=[6, 32, 12, 32, 16, 32, 18])
risks = [
    (1,'報紙廣告未達2則','九(十一)2','每不足1則扣NT$20,000','無上限','報紙刊登截圖/掃描件、版面編號、刊期','J-1報紙廣告'),
    (2,'論壇後新聞稿未達1篇','九(十一)2','扣NT$10,000','NT$10,000','報紙刊登截圖、刊期','J-2活動後新聞稿'),
    (3,'數位廣編稿未達10則','九(十一)3','每不足1則扣NT$10,000','契約價金20%','各則URL截圖、媒體、發布日期','J-3數位廣編稿'),
    (4,'廣編稿總曝光未達200萬人次','九(十一)3','以契約價金20%為限','契約價金20%','後台數據截圖（含日期/觸及）','J-3數位廣編稿'),
    (5,'電視牆/看板未達4,000檔次','九(十一)4','每不足1檔扣NT$1,000','契約價金20%','場域播出紀錄（時段/地點/檔次）','J-4電視牆看板'),
    (6,'網路廣告未達20則','九(十一)5','每不足1則扣NT$1,000','契約價金20%','各則廣告截圖、平台、刊期','J-5網路關鍵字'),
    (7,'長照系列影片未完成','九(十一)6','扣NT$10,000','NT$10,000','影片檔案≥10分鐘、上刊截圖','（M-1，非本表）'),
]
for r, row in enumerate(risks, start=4):
    for c, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = F_BODY
        cell.alignment = Alignment(horizontal='center' if c in (1,3,5) else 'left', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[r].height = 32
ws.freeze_panes = 'A4'

# Save
wb.save('公傳室J類廣告追蹤儀表板.xlsx')
print('OK saved')
